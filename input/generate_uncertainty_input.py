#!/usr/bin/env python
# this script is used to generate input files for uncertainty analysis
# input file includes all optional papameters and a template includes
# unchanged parameters
import os
import pandas as pd
from itertools import product
from openpyxl import Workbook, load_workbook

# range of end_no: [0, 72900]
start_no = 0
end_no = 2
exclude_no = [1]

input_folder = "./uncertainty_input_data"
output_folder = "./scenario_inputs"
def read_para(file_names, sheet=''):
    parameters = {}
    for ix, file_name in enumerate(file_names):
        path = os.path.join(input_folder, file_name)
        sheets = load_workbook(filename=path).worksheets
        df_names = [i.title for i in sheets]
        df_dict = {i:j for i,j in zip(df_names, sheets)}
        if sheet == "":
            parameters[ix] = sheets[0]
        else:
            parameters[ix] = df_dict[sheet]
    return parameters
demand_files = ['demand high.xlsx', 'demand low.xlsx', 'demand medium.xlsx']
discount_factor_files = ['discount factor medium.xlsx']
fuel_price_files = ['fuel price high.xlsx', 'fuel price low.xlsx', 'fuel price medium.xlsx']
inflow_files = ['inflow dry.xlsx', 'inflow normal.xlsx', 'inflow wet.xlsx']

solution_files = ["solution_S7.xlsx", "solution_S8.xlsx", "solution_S9.xlsx", "solution_S10.xlsx", "solution_S11.xlsx", "solution_S12.xlsx", "solution_S13.xlsx", "solution_S14.xlsx", "solution_S15.xlsx", "solution_S16.xlsx"]

technology_upper_bound_files = ['technology upper bound high.xlsx', 'technology upper bound low.xlsx', 'technology upper bound medium.xlsx']

technology_fix_cost_files = ['technology fix cost high.xlsx', 'technology fix cost low.xlsx', 'technology fix cost medium.xlsx']
technology_investment_cost_files = ['technology investment cost high.xlsx', 'technology investment cost low.xlsx', 'technology investment cost medium.xlsx']
technology_variable_cost_files = ['technology variable cost high.xlsx', 'technology variable cost low.xlsx', 'technology variable cost medium.xlsx']

transline_fix_cost_files = ['transline fix cost high.xlsx', 'transline fix cost low.xlsx', 'transline fix cost medium.xlsx']
transline_investment_cost_files = ['transline investment cost high.xlsx', 'transline investment cost low.xlsx', 'transline investment cost medium.xlsx']
transline_variable_cost_files = ['transline variable cost high.xlsx', 'transline variable cost low.xlsx', 'transline variable cost medium.xlsx']
demand = read_para(demand_files)
discount_factor = read_para(discount_factor_files)
fuel_price = read_para(fuel_price_files)
inflow = read_para(inflow_files)
carbon = read_para(solution_files, sheet="carbon")
trans_limit = read_para(solution_files, sheet="trans_limit")
import_limit = read_para(solution_files, sheet="import_limit")
portfolio = read_para(solution_files, sheet="portfolio")
technology_upper_bound = read_para(technology_upper_bound_files)
technology_fix_cost = read_para(technology_fix_cost_files)
technology_investment_cost = read_para(technology_investment_cost_files)
technology_variable_cost = read_para(technology_variable_cost_files)
transline_fix_cost = read_para(transline_fix_cost_files)
transline_variable_cost = read_para(transline_variable_cost_files)
transline_investment_cost = read_para(transline_investment_cost_files)

template = os.path.join(input_folder, "mekong_2020_288h_4_years_AD110dams_GHG1.xlsx")
template_data = load_workbook(filename=template).worksheets
template_dict = {data.title : data for data in template_data}

df_scenarios = pd.DataFrame(columns=
        ["demand" , "discount factor", "fuel price", "inflow",
        "carbon", "trans_limit", "import_limit",
        "portfolio", "technology upper bound",
        "technology fix cost",
        "technology investment cost",
        "technology variable cost",
        "transline fix cost",
        "transline investment cost",
        "transline variable cost"]
)
no = start_no
indices = []
for ix in product(
    range(len(demand_files)), range(len(discount_factor_files)), range(len(fuel_price_files)),
    range(len(inflow_files)), range(len(solution_files)),
    range(len(technology_upper_bound_files)),
    range(len(technology_fix_cost_files)), range(len(transline_fix_cost_files))
):
    # if True:
    if no not in exclude_no:
        # File paths
        update_paras = {
            "demand" : demand[ix[0]],
            "discount factor": discount_factor[ix[1]],
            "fuel price": fuel_price[ix[2]],
            "inflow": inflow[ix[3]],
            "carbon": carbon[ix[4]],
            "trans_limit": trans_limit[ix[4]],
            "import_limit": import_limit[ix[4]],
            "portfolio": portfolio[ix[4]],
            "technology upper bound": technology_upper_bound[ix[5]],
            "technology fix cost": technology_fix_cost[ix[6]],
            "technology investment cost": technology_investment_cost[ix[6]],
            "technology variable cost": technology_variable_cost[ix[6]],
            "transline fix cost": transline_fix_cost[ix[7]],
            "transline investment cost": transline_investment_cost[ix[7]],
            "transline variable cost": transline_variable_cost[ix[7]],
        }
        file_names = {
            "demand": demand_files[ix[0]],
            "discount factor": discount_factor_files[ix[1]],
            "fuel price": fuel_price_files[ix[2]],
            "inflow": inflow_files[ix[3]],
            "carbon": solution_files[ix[4]],
            "trans_limit": solution_files[ix[4]],
            "import_limit": solution_files[ix[4]],
            "portfolio": solution_files[ix[4]],
            "technology upper bound": technology_upper_bound_files[ix[5]],
            "technology fix cost": technology_fix_cost_files[ix[6]],
            "technology investment cost": technology_investment_cost_files[ix[6]],
            "technology variable cost": technology_variable_cost_files[ix[6]],
            "transline fix cost": transline_fix_cost_files[ix[7]],
            "transline investment cost": transline_investment_cost_files[ix[7]],
            "transline variable cost": transline_variable_cost_files[ix[7]]
        }
        # print(update_paras)
        template_data_copy = template_dict.copy()
        for key, val in update_paras.items():
            template_data_copy[key] = val
        new_wb = Workbook()
        new_wb.remove(new_wb.active)
        # Copy each sheet from template_dict into the new workbook
        for sheet_name, sheet in template_data_copy.items():
            new_ws = new_wb.create_sheet(title=sheet_name)  # Create a new sheet
            for row in sheet.iter_rows():
                for cell in row:
                    new_ws[cell.coordinate] = cell.value  # Copy values
        indices.append(no)
        df_scenarios = pd.concat([df_scenarios, pd.DataFrame([file_names])], ignore_index=True)
#         # Replace the demand sheet in the template data
#         for key, val in update_paras.items():
#                 template_data[key] = val

#         # Optionally, save the modified template for this iteration
        output_path = os.path.join(output_folder, f"input_{no}.xlsx")
        new_wb.save(output_path)
#         with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
#             for sheet_name, df in template_data.items():
#                 df.to_excel(writer, sheet_name=sheet_name, index=False)
    no += 1
    df_scenarios.index = indices
    if no >= end_no:
        output_path = os.path.join(output_folder, f"uncertainty_scenarios_{start_no}_{end_no}.xlsx")
        df_scenarios.to_excel(output_path)
        break
